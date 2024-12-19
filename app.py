import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import NoSuchElementException, ElementNotInteractableException, ElementNotSelectableException, TimeoutException, WebDriverException
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import openpyxl
import schedule
from datetime import date, datetime
from time import sleep
import os

stop_schedule = False

# Intervalo de execução do código em MINUTOS
time_exe = 30

# Link do produto 
url = 'https://www.amazon.com.br/HD-SSD-KINGSTON-SA400S37-480GB/dp/B075BKXSCQ/ref=sr_1_3?crid=3ABFVIMZS2O0T&dib=eyJ2IjoiMSJ9.xtLqN2YY2lmra89PEqy7G0Y84YkphyyUWV5twVivkNse1ODjkVlE7xfYt0-FIma8U9i0nv0Se2nMSPd-hyWG2Teo__6mPZ9JRy8ISaS7yBPhkUYWYTAJopOQ4hRMxZ7dzQoHYb3lI3LHfw_YRtgQrv4Fwxhs4tWHKz4EMr7VTkiSDRSWgvk3N6BZT1FtUAJMQc6JscsuzwuHCoTXPfmZm36OgM4cVt-aZFb9XNsALKZBTBNEIumTM7NwH3bVuyT-z9NAVFNpSltZOaTQnUckcRMTscf3YKI2hCTLpSqWQig.WagU7fhpIkWdmNgtfShtJccPMhTWlh9y46wTOG-nbZY&dib_tag=se&keywords=ssd+500gb&qid=1732729870&sprefix=ssd%2Caps%2C160&sr=8-3&ufe=app_do%3Aamzn1.fos.6a09f7ec-d911-4889-ad70-de8dd83c8a74'

def config_driver():
    drive_options = webdriver.ChromeOptions()
    arguments = ['--window-position=200,100', 
                 '--lang=pt-br', 
                 '--incognito', 
                 '--disable-site-isolation-trials',
                 '--disable-blink-features=AutomationControlled',
                 '--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/89.0.4389.82 Safari/537.36'
                 ]

    for argument in arguments:
        drive_options.add_argument(argument)

    drive = webdriver.Chrome(options=drive_options)

    wait = WebDriverWait(
        drive,
        timeout=10,
        poll_frequency=2,
        ignored_exceptions=[NoSuchElementException, ElementNotInteractableException, ElementNotSelectableException]
    )
    return wait, drive

def get_information(wait, driver):
    global url

    print('Acessando o site...')

    try:
        driver.get(url)
        sleep(1)
    except TimeoutException as e:
        print(f'Erro ao acessar o site{url}, {e}')
        raise e
    except WebDriverException as e:
        print(f'Erro inesperado ao acessar o site{url}, {e}')
        raise e
        
    
    print('Pegando informações...')
    try:
        wait.until(EC.presence_of_element_located((By.XPATH, '//span[@id="productTitle"]')))
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="corePriceDisplay_desktop_feature_div"]/div[1]/span[2]/span[2]/span[@class="a-price-whole"]')))
        wait.until(EC.presence_of_element_located((By.XPATH, '//*[@id="corePriceDisplay_desktop_feature_div"]/div[1]/span[2]/span[2]/span[@class="a-price-fraction"]')))

        product_name = driver.find_element(By.XPATH, '//span[@id="productTitle"]')
        product_price1 = driver.find_element(By.XPATH, '//*[@id="corePriceDisplay_desktop_feature_div"]/div[1]/span[2]/span[2]/span[@class="a-price-whole"]')
        product_price2 = driver.find_element(By.XPATH, '//*[@id="corePriceDisplay_desktop_feature_div"]/div[1]/span[2]/span[2]/span[@class="a-price-fraction"]')

        product_name_text = product_name.text
        product_price_text = product_price1.text + ',' + product_price2.text

        return product_name_text, product_price_text, url
    
    except NoSuchElementException as e:
        print(f"Elemento não encontrado na página: {e}")
        raise e
    except Exception as e:
        print(f"Erro ao pegar informações do site: {e}")
        raise e
    
def save_excel(product_name_text, product_price_text, url, date):
    try:
        # Nome do arquivo Excel
        file_name = 'Planilha_de_preços.xlsx'

        # Verifica se o arquivo já existe
        if os.path.exists(file_name):
            # Abre o arquivo existente
            print('Abrindo planilha existente...')
            book = openpyxl.load_workbook(file_name)
        else:
            # Cria um novo arquivo Excel
            print('Criando nova planilha...')
            book = openpyxl.Workbook()

        # Verifica se a aba "Produto" já existe
        if 'Produto' in book.sheetnames:
            # Seleciona a aba
            produto_page = book['Produto']
        else:
            # Cria a aba "Produto"
            produto_page = book.create_sheet('Produto')
            # Adiciona cabeçalhos na aba recém-criada
            produto_page.append(['Produto', 'Preço', 'Data/Horário', 'Link'])

        # Adiciona os dados na próxima linha disponível
        produto_page.append([product_name_text, product_price_text, date, url])

        # Salva as alterações
        book.save(file_name)
        print(f'Dados salvos com sucesso em "{file_name}".')
    except PermissionError:
        print('Erro de permissão: o arquivo Excel está aberto. Feche-o e tente novamente.')
        raise
    except Exception as e:
        print(f'Erro ao criar/salvar a planilha: {e}')
        raise

def get_date():
    #Horario\Data formatado
    now_hour = datetime.now().strftime("%H:%M:%S")
    now_date = date.today().strftime("%d/%m/%Y")
    return f'{now_date} - {now_hour}'

def exe_script():
    global stop_schedule
    wait, driver = config_driver()

    try:

        product_name_text, product_price_text, url = get_information(wait, driver)
        if product_name_text and product_price_text:
            date = get_date()
            save_excel(product_name_text, product_price_text, url, date)
        else:
            print('Informações do produto não foram obtidas')
            stop_schedule = True

    except Exception as e:
        print(f'Erro inesperado na execução do script: {e}')
        stop_schedule = True
    finally:
        driver.quit()
    print('Driver encerrado...')

#Agendando a execução para os próximos 30 mins
def to_schedule():
    global stop_schedule
    if stop_schedule == False:
        try:
            schedule.every(time_exe).minutes.do(exe_script)
            print(f'Agendamento concluído. A tarefa será executada a cada {time_exe} minutos.\nHorário atual: {datetime.now().strftime("%H:%M:%S")}')

            while not stop_schedule:
                schedule.run_pending()
                next_run = schedule.next_run()  # Obtém a hora da próxima execução
                print(f'Próxima execução agendada para: {next_run.strftime("%d/%m/%Y %H:%M:%S")}')
                sleep(900) # Verifica a cada 15 min pra evitar sobrecarga

        except KeyboardInterrupt:
            print("Agendamento interrompido pelo usuário.")
        except Exception as e:
            print(f"Erro no loop de agendamento: {e}")
    else:
        return None

if __name__ == '__main__':
    exe_script()
    to_schedule()